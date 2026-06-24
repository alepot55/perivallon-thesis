#!/usr/bin/env python3
"""
Reverse-sync: notes/<id>.md frontmatter → SOTA_DLxSAT.xlsx

Use after correcting metadata in notes/. Matches Excel rows by title prefix
(first 50 chars, fuzzy lowercase), then overwrites Titolo/Anno/Link if note
has authoritative values. Preserves all other columns and team-added rows.

Claude maintains the Excel via this script so the user never has to edit it
manually.
"""
from __future__ import annotations
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
EXCEL = ROOT / "SOTA_DLxSAT.xlsx"
NOTES = ROOT / "notes"


def parse_frontmatter(text: str) -> dict:
    m = re.match(r"^---\n(.*?)\n---\n", text, re.DOTALL)
    if not m:
        return {}
    out = {}
    for line in m.group(1).split("\n"):
        if ":" not in line:
            continue
        k, _, v = line.partition(":")
        v = v.strip()
        if v == "null":
            out[k.strip()] = None
        elif v.startswith("[") and v.endswith("]"):
            inner = v[1:-1].strip()
            out[k.strip()] = (
                [x.strip().strip('"') for x in inner.split(",") if x.strip()]
                if inner else []
            )
        else:
            try:
                if v.lstrip("-").isdigit():
                    out[k.strip()] = int(v)
                else:
                    out[k.strip()] = v.strip('"')
            except ValueError:
                out[k.strip()] = v.strip('"')
    return out


def find_row(ws, title_prefix: str) -> int | None:
    pl = title_prefix.lower()[:50]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell = row[0]
        if cell.value and pl in str(cell.value).lower():
            return i
    return None


def main():
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Literature Review"]

    updated = 0
    for note in sorted(NOTES.glob("*.md")):
        fm = parse_frontmatter(note.read_text())
        title = fm.get("title")
        if not title:
            continue
        r = find_row(ws, title)
        if not r:
            continue
        # Authoritative link: arxiv ID or DOI URL
        link = None
        if fm.get("arxiv"):
            link = f"arXiv:{fm['arxiv']}"
        elif fm.get("doi"):
            link = f"https://doi.org/{fm['doi']}"
        if link:
            ws.cell(row=r, column=8).value = link
        if fm.get("year"):
            ws.cell(row=r, column=2).value = fm["year"]
        updated += 1

    wb.save(EXCEL)
    print(f"Synced {updated} rows from notes/ → {EXCEL.name}")


if __name__ == "__main__":
    main()
